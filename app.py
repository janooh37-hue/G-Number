import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

FILE_PATH = 'data.xlsx'
SHEET_NAME = 'Sheet1'

class AttendanceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Employee Attendance")
        self.root.geometry("1200x600")
        self.root.configure(bg="#ecf0f1")
        
        self.df = pd.read_excel(FILE_PATH, sheet_name=SHEET_NAME, header=None)
        self.load_employees()
        
        self.create_widgets()
    
    def load_employees(self):
        self.employees = {}
        for idx, row in self.df.iterrows():
            if idx >= 5:
                g_num = str(row[1]).strip() if pd.notna(row[1]) else ""
                name = str(row[2]).strip() if pd.notna(row[2]) else ""
                if g_num and g_num.startswith('G'):
                    self.employees[g_num] = {"name": name, "row": idx + 1}
    
    def create_widgets(self):
        tk.Label(self.root, text="Employee Attendance Entry", bg="#2c3e50", fg="white", 
                 font=("Arial", 16, "bold"), pady=12).pack(fill=tk.X)
        
        main_frame = tk.Frame(self.root, bg="#ecf0f1")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        left_frame = tk.Frame(main_frame, bg="#ecf0f1", width=320)
        left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        left_frame.pack_propagate(False)
        
        tk.Label(left_frame, text="Employees", bg="#ecf0f1", font=("Arial", 12, "bold")).pack(pady=(0, 5))
        
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(left_frame, textvariable=self.search_var, font=("Arial", 11))
        self.search_entry.pack(fill=tk.X, pady=5)
        self.search_entry.bind('<KeyRelease>', self.on_search)
        
        list_frame = tk.Frame(left_frame, bg="white", bd=1, relief=tk.SOLID, height=150)
        list_frame.pack(fill=tk.X, pady=5)
        list_frame.pack_propagate(False)
        
        self.emp_listbox = tk.Listbox(list_frame, font=("Arial", 10), borderwidth=0, highlightthickness=0, selectbackground="#3498db", selectforeground="white")
        self.emp_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.emp_listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.emp_listbox.config(yscrollcommand=scrollbar.set)
        
        self.emp_listbox.bind('<<ListboxSelect>>', self.on_select_employee)
        
        self.emp_info_label = tk.Label(left_frame, text="Select an employee", font=("Arial", 11, "bold"), 
                                       fg="#7f8c8d", bg="#ecf0f1", wraplength=300)
        self.emp_info_label.pack(pady=10)
        
        tk.Label(left_frame, text="Set Attendance", bg="#ecf0f1", font=("Arial", 12, "bold")).pack(pady=(10, 5))
        
        type_frame = tk.LabelFrame(left_frame, text="Leave Type", bg="#ecf0f1", font=("Arial", 10))
        type_frame.pack(fill=tk.X, pady=5)
        
        self.leave_type = tk.StringVar(value="SL")
        for text, val in [("SL - Sick Leave", "SL"), ("AL - Annual Leave", "AL"), ("AB - Absent", "AB"), ("NG - New Guard", "NG"), ("TR - Training", "TR"), ("R - Resigned/Terminated", "R")]:
            ttk.Radiobutton(type_frame, text=text, variable=self.leave_type, value=val).pack(anchor=tk.W, padx=10, pady=2)
        
        date_frame = tk.LabelFrame(left_frame, text="Date Range", bg="#ecf0f1", font=("Arial", 10))
        date_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(date_frame, text="Start Day:", bg="#ecf0f1").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.start_day_var = tk.StringVar(value=str(datetime.now().day))
        ttk.Entry(date_frame, textvariable=self.start_day_var, width=8).grid(row=0, column=1, padx=5, pady=5)
        
        tk.Label(date_frame, text="Days:", bg="#ecf0f1").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.num_days_var = tk.StringVar(value="1")
        ttk.Entry(date_frame, textvariable=self.num_days_var, width=8).grid(row=1, column=1, padx=5, pady=5)
        
        ttk.Button(left_frame, text="Set Entry", command=self.set_entry).pack(pady=10, fill=tk.X)
        
        self.status_label = tk.Label(left_frame, text="", font=("Arial", 9), fg="green", bg="#ecf0f1")
        self.status_label.pack(pady=5)
        
        self.current_employee = None
        
        right_frame = tk.Frame(main_frame, bg="#ecf0f1")
        right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        tk.Label(right_frame, text="Attendance Calendar (Days 1-31)", bg="#ecf0f1", 
                 font=("Arial", 12, "bold")).pack(pady=(0, 10))
        
        calendar_frame = tk.Frame(right_frame, bg="#bdc3c7")
        calendar_frame.pack(fill=tk.BOTH, expand=True)
        
        canvas = tk.Canvas(calendar_frame, bg="#bdc3c7", highlightthickness=0)
        h_scroll = ttk.Scrollbar(calendar_frame, orient="horizontal", command=canvas.xview)
        self.scroll_frame = tk.Frame(canvas, bg="#bdc3c7")
        
        self.scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        
        canvas.create_window((0, 0), window=self.scroll_frame, anchor="nw")
        canvas.configure(xscrollcommand=h_scroll.set)
        
        canvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        h_scroll.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.day_widgets = {}
        self.create_days_horizontal()
        
        self.populate_employee_list()
    
    def create_days_horizontal(self):
        container = tk.Frame(self.scroll_frame, bg="#bdc3c7")
        container.pack()
        
        default_color_map = {"P": "#27ae60", "SL": "#e74c3c", "AL": "#f39c12", "AB": "#9b59b6", "NG": "#3498db", "TR": "#1abc9c", "R": "#7f8c8d"}
        
        self.day_widgets = {}
        
        for day in range(1, 32):
            col_frame = tk.Frame(container, bg="#34495e")
            col_frame.grid(row=0, column=day-1, padx=1, pady=1)
            
            tk.Label(col_frame, text=f"{day}", bg="#34495e", fg="white", 
                    font=("Arial", 8, "bold"), width=4).pack(pady=(2,1))
            
            status_lbl = tk.Label(col_frame, text="P", bg=default_color_map["P"], fg="white", 
                                 font=("Arial", 9, "bold"), width=4, relief=tk.RAISED)
            status_lbl.pack(pady=(0,2))
            self.day_widgets[day] = status_lbl
    
    def populate_employee_list(self, query=""):
        self.emp_listbox.delete(0, tk.END)
        
        sorted_emps = sorted(self.employees.keys())
        count = 0
        for g_num in sorted_emps:
            name = self.employees[g_num]["name"]
            display = f"{g_num} - {name}"
            
            if not query or query.upper() in display.upper():
                self.emp_listbox.insert(tk.END, display)
                count += 1
                if count >= 5 and not query:
                    break
        
        if not query and len(sorted_emps) > 5:
            for g_num in sorted_emps[5:]:
                self.emp_listbox.insert(tk.END, f"{g_num} - {self.employees[g_num]['name']}")
    
    def on_search(self, event=None):
        query = self.search_var.get().strip()
        self.populate_employee_list(query)
    
    def on_select_employee(self, event=None):
        selection = self.emp_listbox.curselection()
        if selection:
            selected = self.emp_listbox.get(selection[0])
            g_num = selected.split(" - ")[0]
            self.current_employee = g_num
            self.emp_info_label.config(text=f"{self.employees[g_num]['name']}\n({g_num})", fg="#2980b9")
            self.load_attendance()
    
    def get_day_column(self, day):
        return 6 + (day - 1)
    
    def load_attendance(self):
        if not self.current_employee:
            return
        
        g_num = self.current_employee
        row_num = self.employees[g_num]["row"]
        
        try:
            wb = load_workbook(FILE_PATH, data_only=True)
            ws = wb[SHEET_NAME]
            
            for day in range(1, 32):
                col_num = self.get_day_column(day)
                value = ws.cell(row=row_num, column=col_num).value
                display = str(value).strip() if value else "P"
                
                color_map = {"P": "#27ae60", "SL": "#e74c3c", "AL": "#f39c12", "AB": "#9b59b6", "NG": "#3498db", "TR": "#1abc9c", "R": "#7f8c8d"}
                bg_color = color_map.get(display, "white")
                fg_color = "white" if display != "P" else "black"
                
                self.day_widgets[day].config(text=display, bg=bg_color, fg=fg_color)
            
            wb.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load: {str(e)}")
    
    def set_entry(self):
        if not self.current_employee:
            messagebox.showerror("Error", "Select an employee first!")
            return
        
        try:
            start_day = int(self.start_day_var.get())
            num_days = int(self.num_days_var.get())
            if start_day < 1 or start_day > 31 or num_days < 1:
                messagebox.showerror("Error", "Invalid day or number!")
                return
        except ValueError:
            messagebox.showerror("Error", "Invalid input!")
            return
        
        leave_type = self.leave_type.get()
        g_num = self.current_employee
        row_num = self.employees[g_num]["row"]
        
        try:
            wb = load_workbook(FILE_PATH)
            ws = wb[SHEET_NAME]
            
            updated_days = []
            for i in range(num_days):
                day = start_day + i
                if day > 31:
                    break
                col_num = self.get_day_column(day)
                ws.cell(row=row_num, column=col_num, value=leave_type)
                updated_days.append(str(day))
            
            wb.save(FILE_PATH)
            wb.close()
            
            self.status_label.config(text=f"✓ Set {leave_type} for days: {', '.join(updated_days)}")
            self.load_attendance()
        except PermissionError:
            messagebox.showerror("Error", "Close data.xlsx in Excel first!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = AttendanceApp(root)
    root.mainloop()
