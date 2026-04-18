import sys
import re
from copy import copy as style_copy
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                                QHBoxLayout, QLabel, QListWidget, QListWidgetItem, QLineEdit, QRadioButton,
                                QPushButton, QFrame, QGroupBox, QScrollArea, QGridLayout, QMessageBox,
                                QComboBox, QMenu, QSpinBox, QSizePolicy)
from PySide6.QtCore import Qt, QEventLoop, QTimer
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

FILE_PATH = 'data.xlsx'
SHEET_NAME = 'Sheet1'

COLOR_MAP = {
    "P": "#27ae60",
    "SL": "#1abc9c", 
    "AL": "#3498db",
    "AB": "#e74c3c",
    "NG": "#f39c12", 
    "TR": "#9b59b6",
    "-": "#7f8c8d"
}

class AttendanceApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("G-Number")
        self.setGeometry(100, 100, 2100, 950)
        self.current_employee = None
        self.employees = {}
        self.current_day = datetime.now().day
        self.current_mode = "Auto Fill"
        self.num_days_mode = "Auto"
        
        self.load_employees()
        self.setup_ui()
        self.apply_styles()
        self.load_totals()
    
    def load_employees(self):
        self.df = pd.read_excel(FILE_PATH, sheet_name=SHEET_NAME, header=None)
        for idx, row in self.df.iterrows():
            if idx >= 5:
                g_num = str(row[1]).strip() if pd.notna(row[1]) else ""
                name = str(row[2]).strip() if pd.notna(row[2]) else ""
                if g_num and g_num.startswith('G'):
                    self.employees[g_num] = {"name": name, "row": idx + 1}
    
    def setup_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(8)
        main_layout.setContentsMargins(10, 10, 10, 10)

        header = QLabel("Employee Attendance Entry")
        header.setObjectName("header")
        main_layout.addWidget(header)

        content = QHBoxLayout()
        content.setSpacing(10)
        main_layout.addLayout(content, 1)

        # ============ LEFT: EMPLOYEES ============
        left = QWidget()
        left.setFixedWidth(300)
        left_layout = QVBoxLayout(left)
        left_layout.setSpacing(8)
        left_layout.setContentsMargins(0, 0, 0, 0)

        emp_label = QLabel("Employees")
        emp_label.setObjectName("section_title")
        left_layout.addWidget(emp_label)

        self.search = QLineEdit()
        self.search.setPlaceholderText("Search name or G-number...")
        self.search.textChanged.connect(self.on_search)
        left_layout.addWidget(self.search)

        self.emp_list = QListWidget()
        self.emp_list.setSpacing(2)
        self.emp_list.itemClicked.connect(self.on_select_employee)
        left_layout.addWidget(self.emp_list, 1)

        self.emp_info = QLabel("Select an employee")
        self.emp_info.setObjectName("emp_info")
        self.emp_info.setWordWrap(True)
        self.emp_info.setTextFormat(Qt.RichText)
        left_layout.addWidget(self.emp_info)

        self.stats_label = QLabel("")
        self.stats_label.setObjectName("stats_label")
        self.stats_label.setWordWrap(True)
        self.stats_label.setTextFormat(Qt.RichText)
        left_layout.addWidget(self.stats_label)

        content.addWidget(left)

        # ============ MIDDLE: CALENDAR ============
        middle = QWidget()
        middle_layout = QVBoxLayout(middle)
        middle_layout.setSpacing(8)
        middle_layout.setContentsMargins(0, 0, 0, 0)

        self.cal_title = QLabel("Attendance Calendar")
        self.cal_title.setObjectName("cal_title")
        self.cal_title.setAlignment(Qt.AlignCenter)
        middle_layout.addWidget(self.cal_title)

        weekdays_widget = QWidget()
        weekdays_grid = QGridLayout(weekdays_widget)
        weekdays_grid.setSpacing(6)
        weekdays_grid.setContentsMargins(0, 0, 0, 0)
        for i, wd in enumerate(["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]):
            lbl = QLabel(wd)
            lbl.setObjectName("weekday_header")
            lbl.setAlignment(Qt.AlignCenter)
            weekdays_grid.addWidget(lbl, 0, i)
            weekdays_grid.setColumnStretch(i, 1)
        middle_layout.addWidget(weekdays_widget)

        self.month_grid_widget = QWidget()
        self.month_grid_widget.setObjectName("month_grid_widget")
        self.month_grid = QGridLayout(self.month_grid_widget)
        self.month_grid.setSpacing(6)
        self.month_grid.setContentsMargins(0, 0, 0, 0)
        middle_layout.addWidget(self.month_grid_widget, 1)

        self.day_cells = {}
        self.build_month_grid()

        legend_label = QLabel(
            "<span style='color:#27ae60'>●</span> P=Present &nbsp;&nbsp;"
            "<span style='color:#1abc9c'>●</span> SL=Sick Leave &nbsp;&nbsp;"
            "<span style='color:#3498db'>●</span> AL=Annual Leave &nbsp;&nbsp;"
            "<span style='color:#e74c3c'>●</span> AB=Absent &nbsp;&nbsp;"
            "<span style='color:#f39c12'>●</span> NG=New Guard &nbsp;&nbsp;"
            "<span style='color:#9b59b6'>●</span> TR=Training &nbsp;&nbsp;"
            "<span style='color:#7f8c8d'>●</span> -=Resigned"
        )
        legend_label.setObjectName("legend_label")
        legend_label.setAlignment(Qt.AlignCenter)
        legend_label.setWordWrap(True)
        middle_layout.addWidget(legend_label)

        content.addWidget(middle, 1)

        # ============ RIGHT: CONTROLS ============
        right = QWidget()
        right.setFixedWidth(320)
        right_layout = QVBoxLayout(right)
        right_layout.setSpacing(8)
        right_layout.setContentsMargins(0, 0, 0, 0)

        config_group = QGroupBox("Configuration")
        config_layout = QGridLayout()
        config_layout.addWidget(QLabel("Days:"), 0, 0)
        self.days_selector = QComboBox()
        self.days_selector.addItems(["Auto (31)", "28", "29", "30", "31"])
        self.days_selector.setCurrentText("Auto (31)")
        self.days_selector.currentTextChanged.connect(self.on_days_changed)
        config_layout.addWidget(self.days_selector, 0, 1)

        config_layout.addWidget(QLabel("Split at #:"), 1, 0)
        self.split_selector = QSpinBox()
        self.split_selector.setRange(1, 999)
        self.split_selector.setValue(250)
        self.split_selector.setToolTip("Employee number where yellow separator is placed (default 250)")
        config_layout.addWidget(self.split_selector, 1, 1)
        config_group.setLayout(config_layout)
        right_layout.addWidget(config_group)

        type_group = QGroupBox("Leave Type")
        type_layout = QVBoxLayout()
        type_layout.setSpacing(2)
        self.leave_type = "P"
        for text, val in [("P - Present", "P"), ("SL - Sick Leave", "SL"), ("AL - Annual Leave", "AL"),
                          ("AB - Absent", "AB"), ("NG - New Guard", "NG"),
                          ("TR - Training", "TR"), ("- - Resigned/Terminated", "-")]:
            rb = QRadioButton(text)
            rb.toggled.connect(lambda checked, v=val: setattr(self, 'leave_type', v) if checked else None)
            if val == "P":
                rb.setChecked(True)
            type_layout.addWidget(rb)
        type_group.setLayout(type_layout)
        right_layout.addWidget(type_group)

        date_group = QGroupBox("Date Range")
        date_layout = QGridLayout()
        date_layout.addWidget(QLabel("Start Day:"), 0, 0)
        self.start_day = QLineEdit(str(datetime.now().day))
        self.start_day.setFixedWidth(80)
        date_layout.addWidget(self.start_day, 0, 1)
        date_layout.addWidget(QLabel("Days:"), 1, 0)
        self.num_days = QLineEdit("1")
        self.num_days.setFixedWidth(80)
        date_layout.addWidget(self.num_days, 1, 1)
        date_group.setLayout(date_layout)
        right_layout.addWidget(date_group)

        self.set_btn = QPushButton("Set Entry")
        self.set_btn.clicked.connect(self.set_entry)
        right_layout.addWidget(self.set_btn)

        self.auto_present_btn = QPushButton("Run Auto Fill")
        self.auto_present_btn.clicked.connect(self.on_run_clicked)
        right_layout.addWidget(self.auto_present_btn)

        mode_row = QHBoxLayout()
        mode_label = QLabel("Mode:")
        mode_label.setObjectName("mode_label")
        mode_row.addWidget(mode_label)
        self.mode_selector = QComboBox()
        self.mode_selector.addItems(["Auto Fill", "Auto Organize"])
        self.mode_selector.currentTextChanged.connect(self.on_mode_changed)
        mode_row.addWidget(self.mode_selector, 1)
        right_layout.addLayout(mode_row)

        self.status = QLabel("")
        self.status.setObjectName("status")
        self.status.setWordWrap(True)
        right_layout.addWidget(self.status)

        right_layout.addStretch()

        totals_group = QGroupBox("Total Attendance")
        totals_layout = QVBoxLayout()
        self.totals_label = QLabel("")
        self.totals_label.setObjectName("totals_label")
        self.totals_label.setTextFormat(Qt.RichText)
        self.totals_label.setWordWrap(True)
        totals_layout.addWidget(self.totals_label)
        totals_group.setLayout(totals_layout)
        right_layout.addWidget(totals_group)

        content.addWidget(right)

        self.populate_employee_list()
    
    def get_month_info(self):
        try:
            wb = load_workbook(FILE_PATH, data_only=True)
            ws = wb[SHEET_NAME]
            month_cell = ws.cell(row=5, column=6).value
            wb.close()
            month_str = str(month_cell).upper() if month_cell else ""
        except Exception:
            month_str = ""

        month_map = {"JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
                     "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12}
        month = None
        for name, num in month_map.items():
            if name in month_str:
                month = num
                break
        if month is None:
            month = datetime.now().month

        year_match = re.search(r'(\d{4})', month_str)
        year = int(year_match.group(1)) if year_match else datetime.now().year

        days_in_month = self.get_actual_days()
        first_weekday = datetime(year, month, 1).weekday()
        return month, year, days_in_month, first_weekday

    def get_month_name(self, m):
        return ["January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December"][m - 1]

    def build_month_grid(self):
        while self.month_grid.count() > 0:
            item = self.month_grid.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self.day_cells = {}

        month, year, days_in_month, first_weekday = self.get_month_info()
        self.cal_title.setText(f"{self.get_month_name(month)} {year}  ·  {days_in_month} days")

        for day in range(1, days_in_month + 1):
            pos = first_weekday + (day - 1)
            row = pos // 7
            col = pos % 7

            frame = QFrame()
            frame.setObjectName("day_cell")
            frame.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            cell_layout = QVBoxLayout(frame)
            cell_layout.setContentsMargins(6, 6, 6, 6)
            cell_layout.setSpacing(4)

            day_num = QLabel(str(day))
            day_num.setObjectName("day_number")
            day_num.setAlignment(Qt.AlignCenter)
            cell_layout.addWidget(day_num)

            status = QLabel("P")
            status.setObjectName("day_status")
            status.setAlignment(Qt.AlignCenter)
            status.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            cell_layout.addWidget(status, 1)

            self.month_grid.addWidget(frame, row, col)
            self.day_cells[day] = {"frame": frame, "number": day_num, "status": status}

        for c in range(7):
            self.month_grid.setColumnStretch(c, 1)
        max_rows = (first_weekday + days_in_month - 1) // 7 + 1
        for r in range(max_rows):
            self.month_grid.setRowStretch(r, 1)
    
    def apply_styles(self):
        self.setStyleSheet("""
            QMainWindow {
                background-color: #1e1e2e;
            }
            #header {
                background-color: #313244;
                color: #cdd6f4;
                font-size: 18px;
                font-weight: bold;
                padding: 15px;
                qproperty-alignment: AlignCenter;
            }
            #section_title {
                color: #cdd6f4;
                font-size: 14px;
                font-weight: bold;
                padding: 8px 0;
            }
            #days_label {
                color: #cdd6f4;
                font-size: 12px;
                font-weight: bold;
            }
            #mode_label {
                color: #cdd6f4;
                font-size: 12px;
                font-weight: bold;
            }
            QComboBox {
                background-color: #313244;
                color: #cdd6f4;
                border: 1px solid #45475a;
                border-radius: 6px;
                padding: 5px;
                font-size: 12px;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 4px solid transparent;
                border-right: 4px solid transparent;
                border-top: 6px solid #cdd6f4;
                margin-right: 5px;
            }
            QComboBox QAbstractItemView {
                background-color: #313244;
                color: #cdd6f4;
                selection-background-color: #45475a;
                border: 1px solid #45475a;
            }
            #legend_label {
                color: #cdd6f4;
                font-size: 12px;
                font-weight: bold;
                padding: 8px;
            }
            #emp_info {
                color: #cdd6f4;
                padding: 12px;
                background-color: #313244;
                border-radius: 8px;
                qproperty-alignment: AlignCenter;
            }
            #stats_label {
                color: #cdd6f4;
                font-size: 12px;
                padding: 10px;
                background-color: #313244;
                border-radius: 8px;
                line-height: 1.6;
            }
            #totals_label {
                color: #cdd6f4;
                font-size: 13px;
                padding: 4px;
                line-height: 1.8;
            }
            #cal_title {
                color: #cdd6f4;
                font-size: 16px;
                font-weight: bold;
                padding: 8px;
                background-color: #313244;
                border-radius: 8px;
            }
            #weekday_header {
                color: #89b4fa;
                font-size: 13px;
                font-weight: bold;
                padding: 6px;
                background-color: #313244;
                border-radius: 6px;
            }
            #month_grid_widget {
                background-color: #181825;
                border-radius: 8px;
                padding: 6px;
            }
            QLineEdit {
                background-color: #313244;
                color: #cdd6f4;
                border: 1px solid #45475a;
                border-radius: 6px;
                padding: 8px;
                font-size: 12px;
            }
            QLineEdit:focus {
                border: 1px solid #89b4fa;
            }
            QSpinBox {
                background-color: #313244;
                color: #cdd6f4;
                border: 1px solid #45475a;
                border-radius: 6px;
                padding: 5px;
                font-size: 12px;
            }
            QSpinBox::up-button, QSpinBox::down-button {
                background-color: #45475a;
                border: none;
                width: 16px;
            }
            QSpinBox::up-button:hover, QSpinBox::down-button:hover {
                background-color: #585b70;
            }
            QListWidget {
                background-color: #313244;
                color: #cdd6f4;
                border: 1px solid #45475a;
                border-radius: 8px;
                font-size: 14px;
            }
            QListWidget::item {
                padding: 8px 10px;
                border-radius: 4px;
                min-height: 28px;
            }
            QListWidget::item:selected {
                background-color: #cdd6f4;
                color: #1e1e2e;
            }
            QListWidget::item:hover {
                background-color: #45475a;
            }
            QGroupBox {
                color: #cdd6f4;
                font-weight: bold;
                border: 1px solid #45475a;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 0 5px;
            }
            QRadioButton {
                color: #cdd6f4;
                padding: 4px;
            }
            QRadioButton::indicator {
                width: 16px;
                height: 16px;
            }
            QRadioButton::indicator:unchecked {
                border: 2px solid #45475a;
                border-radius: 8px;
                background-color: #313244;
            }
            QRadioButton::indicator:checked {
                border: 2px solid #89b4fa;
                border-radius: 8px;
                background-color: #cdd6f4;
            }
            QPushButton {
                background-color: #cdd6f4;
                color: #1e1e2e;
                border: none;
                border-radius: 8px;
                padding: 12px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #b4befe;
            }
            QPushButton:pressed {
                background-color: #74c7ec;
            }
            #status {
                color: #a6e3a1;
                font-size: 12px;
                qproperty-alignment: AlignCenter;
                padding: 5px;
            }
            #day_cell {
                background-color: #313244;
                border-radius: 8px;
                min-width: 70px;
                min-height: 70px;
            }
            #day_number {
                color: #6c7086;
                font-size: 14px;
                font-weight: bold;
                qproperty-alignment: AlignCenter;
                background-color: #45475a;
                border-radius: 4px;
                padding: 3px;
            }
            #day_status {
                color: black;
                font-size: 20px;
                font-weight: bold;
                qproperty-alignment: AlignCenter;
                background-color: #27ae60;
                border-radius: 6px;
                padding: 4px;
            }
        """)
    
    def populate_employee_list(self, query=""):
        self.emp_list.clear()
        sorted_emps = sorted(self.employees.keys())
        q = query.upper()
        for g_num in sorted_emps:
            name = self.employees[g_num]["name"]
            display = f"{g_num} - {name}"
            if not q or q in display.upper():
                item = QListWidgetItem(display)
                item.setToolTip(f"{name}\n({g_num})")
                self.emp_list.addItem(item)

    def on_search(self, text):
        self.populate_employee_list(text)

    def on_select_employee(self, item):
        g_num = item.text().split(" - ")[0]
        self.current_employee = g_num
        name = self.employees[g_num]['name']
        self.emp_info.setText(
            f"<div style='font-size:17px;font-weight:bold;'>{name}</div>"
            f"<div style='color:#89b4fa;font-size:13px;margin-top:4px;'>{g_num}</div>"
        )
        self.load_attendance()
        self.load_stats(g_num)

    def get_day_column(self, day):
        return 6 + (day - 1)

    def load_attendance(self):
        self.load_totals()
        if not self.current_employee:
            self.clear_calendar()
            return
        self.load_attendance_to_grid(self.current_employee)
        self.load_stats(self.current_employee)

    def load_attendance_to_grid(self, g_num):
        row_num = self.employees[g_num]["row"]
        try:
            wb = load_workbook(FILE_PATH, data_only=True)
            ws = wb[SHEET_NAME]
            for day, widgets in self.day_cells.items():
                col_num = self.get_day_column(day)
                value = ws.cell(row=row_num, column=col_num).value
                display = str(value).strip() if value else "P"

                bg_color = COLOR_MAP.get(display, "#27ae60")
                fg_color = "white" if display != "P" else "black"

                if day < self.current_day:
                    day_border = "border: 2px solid #a6e3a1;"
                    day_bg = "#313244"
                    day_fg = "#a6e3a1"
                elif day == self.current_day:
                    day_border = "border: 2px solid #f9e2af;"
                    day_bg = "#313244"
                    day_fg = "#f9e2af"
                else:
                    day_border = "border: 2px solid #45475a;"
                    day_bg = "#45475a"
                    day_fg = "#6c7086"

                widgets["status"].setText(display)
                widgets["status"].setStyleSheet(f"""
                    color: {fg_color};
                    font-size: 20px;
                    font-weight: bold;
                    qproperty-alignment: AlignCenter;
                    background-color: {bg_color};
                    border-radius: 6px;
                    padding: 4px;
                """)
                widgets["number"].setStyleSheet(f"""
                    color: {day_fg};
                    font-size: 14px;
                    font-weight: bold;
                    qproperty-alignment: AlignCenter;
                    background-color: {day_bg};
                    border-radius: 4px;
                    padding: 3px;
                    {day_border}
                """)
            wb.close()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load: {str(e)}")

    def load_totals(self):
        try:
            wb = load_workbook(FILE_PATH, data_only=True)
            ws = wb[SHEET_NAME]
            last_col = 5 + self.get_actual_days()

            last_row = 5
            for row in range(6, 1000):
                g_num = ws.cell(row=row, column=2).value
                if g_num and str(g_num).strip() and str(g_num).startswith('G'):
                    last_row = row

            counts = {"P": 0, "SL": 0, "AL": 0, "AB": 0, "NG": 0, "TR": 0, "-": 0}
            for row in range(6, last_row + 1):
                for col in range(6, last_col + 1):
                    v = ws.cell(row=row, column=col).value
                    s = str(v).strip() if v else ""
                    if s in counts:
                        counts[s] += 1
            wb.close()

            total = sum(counts.values())
            cP, cSL, cAL = COLOR_MAP["P"], COLOR_MAP["SL"], COLOR_MAP["AL"]
            cAB, cNG, cTR, cD = COLOR_MAP["AB"], COLOR_MAP["NG"], COLOR_MAP["TR"], COLOR_MAP["-"]
            html = (
                f"<div style='font-size:12px;color:#a6adc8;margin-bottom:6px;'>"
                f"<b>{last_row - 5}</b> employees &nbsp; <b>{total}</b> entries</div>"
                f"<span style='color:{cP};'>&#9679; P:</span> <b>{counts['P']}</b> &nbsp; "
                f"<span style='color:{cSL};'>&#9679; SL:</span> <b>{counts['SL']}</b> &nbsp; "
                f"<span style='color:{cAL};'>&#9679; AL:</span> <b>{counts['AL']}</b><br>"
                f"<span style='color:{cAB};'>&#9679; AB:</span> <b>{counts['AB']}</b> &nbsp; "
                f"<span style='color:{cNG};'>&#9679; NG:</span> <b>{counts['NG']}</b> &nbsp; "
                f"<span style='color:{cTR};'>&#9679; TR:</span> <b>{counts['TR']}</b><br>"
                f"<span style='color:{cD};'>&#9679; -:</span> <b>{counts['-']}</b>"
            )
            self.totals_label.setText(html)
        except Exception:
            self.totals_label.setText("<span style='color:#f38ba8'>Unable to load totals</span>")

    def clear_calendar(self):
        for day, widgets in self.day_cells.items():
            widgets["status"].setText("P")
            widgets["status"].setStyleSheet("""
                color: white;
                font-size: 20px;
                font-weight: bold;
                qproperty-alignment: AlignCenter;
                background-color: #45475a;
                border-radius: 6px;
                padding: 4px;
            """)
            widgets["number"].setStyleSheet("""
                color: #6c7086;
                font-size: 14px;
                font-weight: bold;
                qproperty-alignment: AlignCenter;
                background-color: #45475a;
                border-radius: 4px;
                padding: 3px;
                border: 2px solid #45475a;
            """)
    
    def load_stats(self, g_num):
        row_num = self.employees[g_num]["row"]
        try:
            wb = load_workbook(FILE_PATH, data_only=True)
            ws = wb[SHEET_NAME]
            
            counts = {"P": 0, "SL": 0, "AL": 0, "AB": 0, "NG": 0, "TR": 0, "-": 0}
            
            for day in range(1, self.current_day + 1):
                col_num = self.get_day_column(day)
                value = ws.cell(row=row_num, column=col_num).value
                display = str(value).strip() if value else "P"
                if display in counts:
                    counts[display] += 1
            
            stats_text = f"""
<span style='color: {COLOR_MAP["P"]};'>● P:</span> {counts['P']} &nbsp;
<span style='color: {COLOR_MAP["SL"]};'>● SL:</span> {counts['SL']} &nbsp;
<span style='color: {COLOR_MAP["AL"]};'>● AL:</span> {counts['AL']}<br>
<span style='color: {COLOR_MAP["AB"]};'>● AB:</span> {counts['AB']} &nbsp;
<span style='color: {COLOR_MAP["NG"]};'>● NG:</span> {counts['NG']} &nbsp;
<span style='color: {COLOR_MAP["TR"]};'>● TR:</span> {counts['TR']}
"""
            self.stats_label.setText(stats_text)
            wb.close()
        except Exception as e:
            self.stats_label.setText("")
    
    def set_entry(self):
        if not self.current_employee:
            QMessageBox.warning(self, "Error", "Select an employee first!")
            return
        
        try:
            start_day = int(self.start_day.text())
            num_days = int(self.num_days.text())
            if start_day < 1 or start_day > 31 or num_days < 1:
                raise ValueError()
        except ValueError:
            QMessageBox.warning(self, "Error", "Invalid day or number!")
            return
        
        leave_type = self.leave_type
        g_num = self.current_employee
        row_num = self.employees[g_num]["row"]
        
        try:
            wb = load_workbook(FILE_PATH)
            ws = wb[SHEET_NAME]
            
            updated_days = []
            for i in range(num_days):
                day = start_day + i
                if day > 31:
                    break
                col_num = self.get_day_column(day)
                ws.cell(row=row_num, column=col_num, value=leave_type)
                updated_days.append(str(day))
            
            wb.save(FILE_PATH)
            wb.close()
            
            self.status.setText(f"✓ Set {leave_type} for days: {', '.join(updated_days)}")
            self.load_attendance()
            self.load_stats(g_num)
        except PermissionError:
            QMessageBox.critical(self, "Error", "Close data.xlsx in Excel first!")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed: {str(e)}")
    
    def auto_fill_present(self):
        try:
            wb = load_workbook(FILE_PATH)
            ws = wb[SHEET_NAME]
            
            last_row = 5
            for row in range(6, 1000):
                g_num = ws.cell(row=row, column=2).value
                if g_num and str(g_num).strip() and str(g_num).startswith('G'):
                    last_row = row
            
            if last_row == 5:
                QMessageBox.warning(self, "Error", "No employees found!")
                wb.close()
                return
                
            reply = QMessageBox.question(
                self, "Confirm", 
                f"This will fill all empty cells in columns F to AJ (rows 6-{last_row}) with 'P'. Continue?",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                wb.close()
                return
            
            filled_count = 0
            for row in range(6, last_row + 1):
                for col in range(6, 37):
                    cell = ws.cell(row=row, column=col)
                    if cell.value is None or str(cell.value).strip() == "":
                        cell.value = "P"
                        filled_count += 1
            
            wb.save(FILE_PATH)
            wb.close()
            
            if filled_count > 0:
                QMessageBox.information(self, "Done", f"Added 'P' to {filled_count} cells")
            else:
                QMessageBox.information(self, "Done", "All cells already filled")
            self.load_attendance()
        except PermissionError:
            QMessageBox.critical(self, "Error", "Close data.xlsx in Excel first!")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed: {str(e)}")

    def on_days_changed(self, text):
        if text.startswith("Auto"):
            self.num_days_mode = "Auto"
            actual_days = self.get_actual_days()
            self.days_selector.blockSignals(True)
            self.days_selector.setItemText(0, f"Auto ({actual_days})")
            self.days_selector.blockSignals(False)
        else:
            self.num_days_mode = text
        self.build_month_grid()
        if self.current_employee:
            self.load_attendance_to_grid(self.current_employee)
        else:
            self.clear_calendar()

    def on_mode_changed(self, mode):
        self.current_mode = mode
        self.auto_present_btn.setText(f"Run {mode}")
    
    def on_run_clicked(self):
        if self.current_mode == "Auto Fill":
            self.auto_fill_present()
        else:
            self.auto_organize()
    
    def get_actual_days(self):
        mode = self.num_days_mode
        if mode == "Auto":
            try:
                wb = load_workbook(FILE_PATH, data_only=True)
                ws = wb[SHEET_NAME]
                month_cell = ws.cell(row=5, column=6).value
                month_str = str(month_cell).upper() if month_cell else ""
                wb.close()
                
                if any(m in month_str for m in ["JAN", "MAR", "MAY", "JUL", "AUG", "OCT", "DEC"]):
                    return 31
                elif any(m in month_str for m in ["APR", "JUN", "SEP", "NOV"]):
                    return 30
                elif "FEB" in month_str:
                    year = datetime.now().year
                    if (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0):
                        return 29
                    return 28
                return 31
            except:
                return 31
        return int(mode)
    
    def auto_fill_present(self):
        try:
            wb = load_workbook(FILE_PATH)
            ws = wb[SHEET_NAME]
            
            actual_days = self.get_actual_days()
            last_col = 5 + actual_days
            
            last_row = 5
            for row in range(6, 1000):
                g_num = ws.cell(row=row, column=2).value
                if g_num and str(g_num).strip() and str(g_num).startswith('G'):
                    last_row = row
            
            if last_row == 5:
                QMessageBox.warning(self, "Error", "No employees found!")
                wb.close()
                return
            
            reply = QMessageBox.question(
                self, "Confirm", 
                f"This will fill all empty cells in columns F to column {last_col} (rows 6-{last_row}) with 'P'. Continue?",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                wb.close()
                return
            
            filled_count = 0
            for row in range(6, last_row + 1):
                for col in range(6, last_col + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value is None or str(cell.value).strip() == "":
                        cell.value = "P"
                        filled_count += 1
            
            wb.save(FILE_PATH)
            wb.close()
            
            if filled_count > 0:
                QMessageBox.information(self, "Done", f"Added 'P' to {filled_count} cells")
            else:
                QMessageBox.information(self, "Done", "All cells already filled")
            self.load_attendance()
        except PermissionError:
            QMessageBox.critical(self, "Error", "Close data.xlsx in Excel first!")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed: {str(e)}")
    
    def auto_organize(self):
        try:
            wb = load_workbook(FILE_PATH)
            ws = wb[SHEET_NAME]

            day_count = self.get_actual_days()
            first_day_col = 6
            last_day_col = 5 + day_count
            max_col = 42

            last_emp_row = 5
            for row in range(6, 1000):
                g_num = ws.cell(row=row, column=2).value
                if g_num and str(g_num).strip() and str(g_num).startswith('G'):
                    last_emp_row = row

            if last_emp_row == 5:
                QMessageBox.warning(self, "Error", "No employees found!")
                wb.close()
                return

            split_emp_num = self.split_selector.value()
            emp_split_row = None
            for row in range(6, last_emp_row + 1):
                a_val = ws.cell(row=row, column=1).value
                try:
                    if a_val is not None and int(a_val) == split_emp_num:
                        emp_split_row = row
                        break
                except (TypeError, ValueError):
                    continue

            split_row = emp_split_row if emp_split_row is not None else last_emp_row
            yellow_row1 = split_row + 1
            yellow_row2 = split_row + 2

            ng_rows = set()
            for row in range(6, last_emp_row + 1):
                for col in range(first_day_col, last_day_col + 1):
                    v = ws.cell(row=row, column=col).value
                    if v and str(v).strip().upper() == "NG":
                        ng_rows.add(row)
                        break

            leave_counts = {"SL": 0, "AL": 0, "AB": 0, "TR": 0}
            for row in range(6, last_emp_row + 1):
                if row in ng_rows:
                    continue
                for col in range(first_day_col, last_day_col + 1):
                    v = ws.cell(row=row, column=col).value
                    s = str(v).strip() if v else ""
                    if s in leave_counts:
                        leave_counts[s] += 1

            for row in range(6, last_emp_row + 1):
                if row in ng_rows:
                    continue
                for col in range(first_day_col, last_day_col + 1):
                    v = ws.cell(row=row, column=col).value
                    s = str(v).strip() if v else ""
                    if not s:
                        continue
                    if s == "-":
                        continue
                    ws.cell(row=row, column=col, value="P")

            shift_start = yellow_row1
            shift_end = ws.max_row
            delta = 2
            new_last_emp_row = last_emp_row + delta if shift_start <= last_emp_row else last_emp_row

            if shift_start <= shift_end:
                def shift_formula(formula, d, r_lo, r_hi):
                    if not isinstance(formula, str) or not formula.startswith('='):
                        return formula
                    pat = re.compile(r'(?<!\$)([A-Z]+)(\d+)')
                    def replacer(m):
                        row_num = int(m.group(2))
                        if r_lo <= row_num <= r_hi:
                            row_num += d
                        return m.group(1) + str(row_num)
                    return pat.sub(replacer, formula)

                saved_merges = []
                for mr in list(ws.merged_cells.ranges):
                    if mr.min_row >= shift_start and mr.max_row <= shift_end:
                        saved_merges.append((mr.min_row, mr.min_col, mr.max_row, mr.max_col))
                for min_r, min_c, max_r, max_c in saved_merges:
                    rng = f"{get_column_letter(min_c)}{min_r}:{get_column_letter(max_c)}{max_r}"
                    ws.unmerge_cells(rng)

                saved_cells = {}
                for row in range(shift_start, shift_end + 1):
                    row_height = ws.row_dimensions[row].height
                    row_data = {}
                    for col in range(1, max_col + 1):
                        src = ws.cell(row=row, column=col)
                        row_data[col] = {
                            'value': src.value,
                            'font': style_copy(src.font),
                            'fill': style_copy(src.fill),
                            'border': style_copy(src.border),
                            'alignment': style_copy(src.alignment),
                            'number_format': src.number_format,
                        }
                    saved_cells[row] = (row_data, row_height)

                for row in range(shift_start, shift_end + 1):
                    for col in range(1, max_col + 1):
                        ws.cell(row=row, column=col).value = None

                for old_row, (row_data, height) in saved_cells.items():
                    new_row = old_row + delta
                    if height is not None:
                        ws.row_dimensions[new_row].height = height
                    for col, props in row_data.items():
                        dst = ws.cell(row=new_row, column=col)
                        dst.value = shift_formula(props['value'], delta, shift_start, shift_end)
                        dst.font = props['font']
                        dst.fill = props['fill']
                        dst.border = props['border']
                        dst.alignment = props['alignment']
                        dst.number_format = props['number_format']

                for min_r, min_c, max_r, max_c in saved_merges:
                    new_rng = f"{get_column_letter(min_c)}{min_r + delta}:{get_column_letter(max_c)}{max_r + delta}"
                    ws.merge_cells(new_rng)

                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                for row in [yellow_row1, yellow_row2]:
                    for col in range(1, last_day_col + 1):
                        ws.cell(row=row, column=col).fill = yellow_fill
                    for col in range(37, max_col + 1):
                        ws.cell(row=row, column=col).fill = yellow_fill

                new_data_last = last_emp_row + delta
                cf_obj = ws.conditional_formatting
                cf_rules = cf_obj._cf_rules
                collected = []
                for cf_range, rules in list(cf_rules.items()):
                    new_sqrefs = []
                    for cr in cf_range.sqref.ranges:
                        min_col, min_row, max_col_cr, max_row = cr.bounds
                        if min_row >= shift_start and max_row <= shift_end:
                            min_row += delta
                            max_row += delta
                        elif min_row < shift_start <= max_row:
                            if max_row < new_data_last:
                                max_row = new_data_last
                        new_sqrefs.append(f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col_cr)}{max_row}")
                    collected.append((" ".join(new_sqrefs), list(rules)))
                cf_rules.clear()
                for sqref_str, rules_list in collected:
                    for rule in rules_list:
                        cf_obj.add(sqref_str, rule)

                for r in range(new_data_last + 1, ws.max_row + 1):
                    for c in range(1, 10):
                        v = ws.cell(row=r, column=c).value
                        if isinstance(v, str) and v != v.strip() and v.strip() in ("SL", "AL", "AB", "TR", "-", "S", "R", "P"):
                            ws.cell(row=r, column=c).value = v.strip()

            ng_rows_new = set()
            for row in range(yellow_row2 + 1, new_last_emp_row + 1):
                for col in range(first_day_col, last_day_col + 1):
                    v = ws.cell(row=row, column=col).value
                    if v and str(v).strip().upper() == "NG":
                        ng_rows_new.add(row)
                        break

            cursor_row = yellow_row2 + 1
            cursor_col = first_day_col
            while cursor_row <= new_last_emp_row and cursor_row in ng_rows_new:
                cursor_row += 1

            def advance():
                nonlocal cursor_row, cursor_col
                cursor_col += 1
                if cursor_col > last_day_col:
                    cursor_col = first_day_col
                    cursor_row += 1
                    while cursor_row <= new_last_emp_row and cursor_row in ng_rows_new:
                        cursor_row += 1

            leave_placed = {"SL": 0, "AL": 0, "AB": 0, "TR": 0}
            for leave_type in ["SL", "AL", "AB", "TR"]:
                remaining = leave_counts[leave_type]
                while remaining > 0 and cursor_row <= new_last_emp_row:
                    cell_val = ws.cell(row=cursor_row, column=cursor_col).value
                    s = str(cell_val).strip() if cell_val else ""
                    if s == "P":
                        ws.cell(row=cursor_row, column=cursor_col, value=leave_type)
                        remaining -= 1
                        leave_placed[leave_type] += 1
                    advance()
                if cursor_row > new_last_emp_row:
                    break

            wb.save(FILE_PATH)
            wb.close()

            lost = {k: leave_counts[k] - leave_placed[k] for k in leave_counts}
            msg_lines = [
                f"Collected: SL={leave_counts['SL']}, AL={leave_counts['AL']}, AB={leave_counts['AB']}, TR={leave_counts['TR']}",
                f"Placed:    SL={leave_placed['SL']}, AL={leave_placed['AL']}, AB={leave_placed['AB']}, TR={leave_placed['TR']}",
            ]
            if any(v > 0 for v in lost.values()):
                msg_lines.append(f"Lost (no room): SL={lost['SL']}, AL={lost['AL']}, AB={lost['AB']}, TR={lost['TR']}")
            QMessageBox.information(self, "Done", "Auto Organize complete!\n\n" + "\n".join(msg_lines))
            self.load_attendance()
        except PermissionError:
            QMessageBox.critical(self, "Error", "Close data.xlsx in Excel first!")
        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Error", f"Failed: {str(e)}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = AttendanceApp()
    window.show()
    sys.exit(app.exec())
